import bpy
from . import prop

from collections import defaultdict
import json
import subprocess, os, platform
import time


import pandas as pd
from openpyxl import load_workbook

import zipfile
import xml.parsers.expat

import ifcopenshell
import blenderbim.tool as tool
from blenderbim.bim.ifc import IfcStore
replace_with_IfcStore = "C:\\Algemeen\\07_ifcopenshell\\00_ifc\\02_ifc_library\\IFC Schependomlaan.ifc"
#replace_with_IfcStore ="C:\\Algemeen\\07_ifcopenshell\\00_ifc\\02_ifc_library\\IFC4 demo.ifc"

#1. allow for multiple classification

class Element(list):
    def __init__(self, name, attrs):
        self.name = name
        self.attrs = attrs


class TreeBuilder:
    def __init__(self):
        self.root = Element("root", None)
        self.path = [self.root]

    def start_element(self, name, attrs):
        element = Element(name, attrs)
        self.path[-1].append(element)
        self.path.append(element)

    def end_element(self, name):
        assert name == self.path[-1].name
        self.path.pop()

    def char_data(self, data):
        self.path[-1].append(data)


def get_hidden_rows(node):
    row = 0
    for e in node:
        if not isinstance(e, Element):
            continue
        yield from get_hidden_rows(e)
        if e.name != "table:table-row":
            continue
        attrs = e.attrs
        rows = int(attrs.get("table:number-rows-repeated", 1))
        if "table:visibility" in attrs.keys():  
            yield from range(row, row + rows)
        row += rows


class ConstructDataFrame:
    def __init__(self, context):

        start_time = time.perf_counter()
        
        ifc_dictionary      = defaultdict(list) 
        ifc_properties      = context.scene.ifc_properties
        custom_collection   = context.scene.custom_collection

        #ifc_file = ifcopenshell.open(IfcStore.path)
        ifc_properties.my_spreadsheet_file  = IfcStore.path.replace('.ifc','_blenderbim') 
        ifc_file = ifcopenshell.open(IfcStore.path)
        #ifc_file = ifcopenshell.open(replace_with_IfcStore)
        products = ifc_file.by_type('IfcProduct')

        custom_propertyset_list = []
        common_property_dict = {}
        quantity_property_dict = {}

        wm = bpy.context.window_manager
        total = len(products)
        wm.progress_begin(0, total) 

        for custom_property in custom_collection.items:
            custom_propertyset_list.append(custom_property.name)
        custom_property_unique_list = []

        seen = set()
        for item in custom_propertyset_list:
            if item not in seen:
                seen.add(item)
                custom_property_unique_list.append(item)

        for my_ifcproperty in ifc_properties.__annotations__.keys():
            my_ifcpropertyvalue = getattr(ifc_properties, my_ifcproperty)
            if my_ifcproperty.startswith('my_property'):
                common_property_dict[my_ifcproperty.replace('my_property_','')] = my_ifcpropertyvalue

            if my_ifcproperty.startswith('my_quantity'):
                quantity_property_dict[my_ifcproperty.replace('my_quantity_','')] = my_ifcpropertyvalue 

   
        for i, product in enumerate(products):
            wm.progress_update(i)
            ifc_dictionary[prop.prop_globalid].append(str(product.GlobalId))

            if ifc_properties.my_ifcbuildingstorey:
                ifc_dictionary[prop.prop_ifcbuildingstorey].append(self.get_ifc_building_storey(    context,
                                                                                                    ifc_product=product)[0])

            if ifc_properties.my_ifcproduct:
                ifc_dictionary[prop.prop_ifcproduct].append(str(product.is_a()))
            
            if ifc_properties.my_ifcproductname:
                ifc_dictionary[prop.prop_ifcproductname].append(str(product.Name))

            if ifc_properties.my_ifcproducttypename:
                ifc_dictionary[prop.prop_ifcproducttypename].append(self.get_ifc_type(context,
                                                                                      ifc_product=product)[0])

            if ifc_properties.my_ifcclassification:
                ifc_dictionary[prop.prop_classification].append(self.get_ifc_classification(context,
                                                                                            ifc_product=product)[0])

            if ifc_properties.my_ifcmaterial:
                ifc_dictionary[prop.prop_materials].append(self.get_ifc_materials(context,
                                                                                  ifc_product=product)[0])
                
     
            ifc_pset_common = 'Pset_' +  (str(product.is_a()).replace('Ifc','')) + 'Common'

            for k,v in common_property_dict.items():
                if v:
                    property = str(k)
                    ifc_dictionary[property].append(self.get_ifc_properties_and_quantities( context,
                                                                                            ifc_product=product,
                                                                                            ifc_propertyset_name=ifc_pset_common,
                                                                                            ifc_property_name=property)[0])

            for k,v in quantity_property_dict.items():
                if v:
                    property = str(k)
                    ifc_dictionary[property].append(self.get_ifc_properties_and_quantities( context,
                                                                                            ifc_product=product,
                                                                                            ifc_propertyset_name=prop.prop_basequantities,
                                                                                            ifc_property_name=property)[0])
                        
            if len(custom_collection.items) > 0:
                for item in custom_property_unique_list:
                    property_set = str(item).split('.')[0]
                    property_name = str(item).split('.')[1]
                    #if propertyset meets re conditions, then run dictionry
                    #if property does not meet dictionary then 
                    ifc_dictionary[item].append(str(self.get_ifc_properties_and_quantities(context,
                                                                                        ifc_product=product,
                                                                                        ifc_propertyset_name=property_set,
                                                                                        ifc_property_name=property_name)[0]))
      
        df = pd.DataFrame(ifc_dictionary)
        self.df = df

        print (time.perf_counter() - start_time, "seconds for the dataframe to be created")
        wm.progress_end()

    def get_ifc_building_storey(self, context,ifc_product):

        building_storey_list = []
            
        spatial_container = ifcopenshell.util.element.get_container(ifc_product)
        
        if spatial_container:    
            building_storey_list.append(spatial_container.Name)
            
        if not building_storey_list:
            building_storey_list.append(None)
   
        return tuple(building_storey_list)

    def get_ifc_type(self, context, ifc_product):
    
        ifc_type_list = []
        
        if ifc_product: 
            ifcproduct_type = ifcopenshell.util.element.get_type(ifc_product)
    
            if ifcproduct_type:
                ifc_type_list.append(ifcproduct_type.Name)
        
        if not ifc_type_list:
            ifc_type_list.append(None)

        return tuple(ifc_type_list)
     

    def get_ifc_classification(self, context, ifc_product):
    
        classification_list = []

        # Elements may have multiple classification references assigned
        references = ifcopenshell.util.classification.get_references(ifc_product)
        
        if ifc_product:
            for reference in references:
                system = ifcopenshell.util.classification.get_classification(reference)
                classification_list.append(str(system.Name) + '\n' + str(reference[1]) +  '\n' + str(reference[2]))
                     
        if not classification_list:
            classification_list.append(None)  
            
        return tuple(classification_list)

    def get_ifc_materials(self, context, ifc_product):
    
        material_list = []
        
        if ifc_product:
            ifc_material = ifcopenshell.util.element.get_material(ifc_product)
            if ifc_material:
                
                if ifc_material.is_a('IfcMaterial'):
                    material_list.append(ifc_material.Name)
                
                if ifc_material.is_a('IfcMaterialList'):
                    for materials in ifc_material.Materials:
                        material_list.append(materials.Name)
                
                if ifc_material.is_a('IfcMaterialConstituentSet'):
                    for material_constituents in ifc_material.MaterialConstituents:
                        material_list.append(material_constituents.Material.Name)
                
                if ifc_material.is_a('IfcMaterialLayerSetUsage'):
                    for material_layer in ifc_material.ForLayerSet.MaterialLayers:
                        material_list.append(material_layer.Material.Name)
                    
                if ifc_material.is_a('IfcMaterialProfileSetUsage'):
                    for material_profile in (ifc_material.ForProfileSet.MaterialProfiles):
                        material_list.append(material_profile.Material.Name)
        
        if not material_list:
            material_list.append('None')

        joined_material_list = '\n'.join(material_list)
            
        return [joined_material_list]

    def get_ifc_properties_and_quantities(self, context, ifc_product, ifc_propertyset_name, ifc_property_name):

        ifc_property_list = []

        if ifc_product:
            pset_name = 'Pset_' +  (str(ifc_product.is_a()).replace('Ifc','')) + 'Common'
            ifc_property_list.append(str(ifcopenshell.util.element.get_pset(ifc_product, ifc_propertyset_name, ifc_property_name)))
                    
        if not ifc_property_list:
            ifc_property_list.append(None)
   
        return ifc_property_list


class ExportToSpreadSheet(bpy.types.Operator):
    """Export to a .xlsx or .ods file"""
    bl_idname = "export.tospreadsheet"
    bl_label = "Create spreadsheet"

    def execute(self, context):

        self.create_spreadsheet(context)
         
        return {'FINISHED'}

    
    def create_spreadsheet(self, context):

        ifc_properties = context.scene.ifc_properties

        if ifc_properties.ods_or_xlsx == 'XLSX':
            ifc_properties.my_spreadsheetfile =''

            construct_data_frame = ConstructDataFrame(context)
            #spreadsheet_filepath = replace_with_IfcStore.replace('.ifc','_blenderbim.xlsx')
            #IfcStore.path.replace('.ifc','_blenderbim.xlsx')
            #ifc_properties.my_spreadsheetfile + '.xlsx'
            #spreadsheet_filepath    = ifc_properties.my_spreadsheetfile# + '.xlsx'

            spreadsheet_filepath    = str(IfcStore.path).replace('.ifc','.xlsx')

            writer = pd.ExcelWriter(spreadsheet_filepath, engine='xlsxwriter')
            construct_data_frame.df.to_excel(writer, sheet_name=prop.prop_workbook, startrow=1, header=False, index=False)
            worksheet = writer.sheets[prop.prop_workbook]

            (max_row, max_col) = construct_data_frame.df.shape
        
            column_settings = []
            for header in construct_data_frame.df.columns:
                column_settings.append({'header': header})
      
            worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})
            worksheet.set_column(0, max_col - 1, 30)

            ifc_properties.my_spreadsheetfile = spreadsheet_filepath
            writer.close()

            print ("Spreadsheet is created at: ", spreadsheet_filepath)
            self.open_file_on_each_os(spreadsheet_filepath=spreadsheet_filepath)

        if ifc_properties.ods_or_xlsx == 'ODS':

            ifc_properties.my_spreadsheetfile =''
            
            construct_data_frame = ConstructDataFrame(context)
    
            #spreadsheet_filepath    = replace_with_IfcStore.replace('.ifc','_blenderbim.ods')
            spreadsheet_filepath    = str(IfcStore.path).replace('.ifc','.ods')
            writer                  = pd.ExcelWriter(spreadsheet_filepath, engine='odf')
            construct_data_frame.df.to_excel(writer, sheet_name=prop.prop_workbook, startrow=0, header=True, index=False)
            worksheet               = writer.sheets[prop.prop_workbook]
            writer.close()

            ifc_properties.my_spreadsheetfile = spreadsheet_filepath
            print ("Spreadsheet is created at: ", spreadsheet_filepath)
            self.open_file_on_each_os(spreadsheet_filepath=spreadsheet_filepath)

    def open_file_on_each_os(self, spreadsheet_filepath):

        if platform.system() == 'Darwin':       # macOS
            subprocess.call(('open', spreadsheet_filepath))
        elif platform.system() == 'Windows':    # Windows
            os.startfile(spreadsheet_filepath)
        else:                                   # linux variants
            subprocess.call(('xdg-open', spreadsheet_filepath))

class FilterIFCElements(bpy.types.Operator):
    """Show the IFC elements you filtered in the spreadsheet"""
    bl_idname = "object.filter_ifc_elements"
    bl_label = "select the IFC elements"
    

    def execute(self, context):
   
        ifc_properties = context.scene.ifc_properties 
         
        if ifc_properties.my_spreadsheetfile:
            if ifc_properties.my_spreadsheetfile.endswith(".xlsx"):
                
                workbook_openpyxl = load_workbook(ifc_properties.my_spreadsheetfile)
                worksheet_openpyxl = workbook_openpyxl[prop.prop_workbook] 
                
                global_id_filtered_list = []
                       
                for row_idx in range(3, worksheet_openpyxl.max_row + 1):
                    if not worksheet_openpyxl.row_dimensions[row_idx].hidden:
                        cell = worksheet_openpyxl[f"A{row_idx}"]
                        global_id_filtered_list.append(str(cell.value))

                self.filter_IFC_elements(context, guid_list=global_id_filtered_list)
                
                return {'FINISHED'}

               
            if ifc_properties.my_spreadsheetfile.endswith(".ods"):
                print ("Retrieving data from: " , ifc_properties.my_spreadsheetfile)
            
                ziparchive = zipfile.ZipFile(ifc_properties.my_spreadsheetfile, "r")
                xmldata = ziparchive.read("content.xml")
                ziparchive.close()
                
                parser = xml.parsers.expat.ParserCreate()
                treebuilder = TreeBuilder()
                parser.StartElementHandler  = treebuilder.start_element
                parser.EndElementHandler    = treebuilder.end_element
                parser.CharacterDataHandler = treebuilder.char_data
                parser.Parse(xmldata, True)

                hidden_rows = get_hidden_rows(treebuilder.root)
                dataframe = pd.read_excel(ifc_properties.my_spreadsheetfile, sheet_name=prop.prop_workbook, skiprows=hidden_rows, engine="odf")
                self.filter_IFC_elements(context, guid_list=dataframe['GlobalId'].values.tolist())
                
                return {'FINISHED'}
               
                
    def filter_IFC_elements(self, context, guid_list):
        
        print ("Filtering IFC elements")
        
        outliner = next(a for a in bpy.context.screen.areas if a.type == "OUTLINER") 
        outliner.spaces[0].show_restrict_column_viewport = not outliner.spaces[0].show_restrict_column_viewport
        
        bpy.ops.object.select_all(action='DESELECT')
      
        for obj in bpy.context.view_layer.objects:
            element = tool.Ifc.get_entity(obj)
            if element is None:        
                obj.hide_viewport = True
                continue
            data = element.get_info()
            obj.hide_viewport = data.get("GlobalId", False) not in guid_list

        bpy.ops.object.select_all(action='SELECT') 
        
       
class UnhideIFCElements(bpy.types.Operator):
    """Unhide all IFC elements"""
    bl_idname = "object.unhide_all"
    bl_label = "Unhide All"

    def execute(self, context):
        for obj in bpy.data.objects:
            obj.hide_viewport = False 
        
        return {'FINISHED'}  


class CustomCollectionActions(bpy.types.Operator):
    bl_idname = "custom.collection_actions"
    bl_label = "Execute"
    action: bpy.props.EnumProperty(items=(("add",) * 3,("remove",) * 3,),)
    index: bpy.props.IntProperty(default=-1)

    def execute(self, context):

        custom_collection = context.scene.custom_collection

        if self.action == "add":        
            custom_item = custom_collection.items.add()  
 
        if self.action == "remove":
            custom_collection.items.remove(self.index)

        return {"FINISHED"}  

    def set_configuration(context,property_name):

        custom_collection = context.scene.custom_collection  
        custom_collection.items.add().name = property_name
     
        return {"FINISHED"}        


class SaveAndLoadSelection(bpy.types.Operator):
    bl_idname = "save.save_and_load_selection"
    bl_label = "Save selection"

    def execute(self, context):

        configuration_dictionary = {}
        ifc_properties = context.scene.ifc_properties
        custom_items = context.scene.custom_collection.items
        
        for my_ifcproperty in ifc_properties.__annotations__.keys():
            my_ifcpropertyvalue = getattr(ifc_properties, my_ifcproperty)
            configuration_dictionary[my_ifcproperty] = my_ifcpropertyvalue

        for prop_name_custom in custom_items.keys():
            prop_value_custom = custom_items[prop_name_custom]
            configuration_dictionary['my_ifccustomproperty' + prop_value_custom.name] = prop_value_custom.name

        with open(replace_with_IfcStore.replace('.ifc','_selectionset.json'), "w") as selection_file:
            json.dump(configuration_dictionary, selection_file, indent=4)

        ifc_properties.my_selectionload = str(replace_with_IfcStore.replace('.ifc','_selectionset.json'))
        print ('Selection has been saved at: ', ifc_properties.my_selectionload)

        return {"FINISHED"} 

class ConfirmSelection(bpy.types.Operator):
    bl_idname = "save.confirm_selection"
    bl_label = "Confirm Selection"

    def execute(self, context):

        ifc_properties = context.scene.ifc_properties
        custom_collections_actions = CustomCollectionActions
        custom_collection = context.scene.custom_collection
        set_configuration = custom_collections_actions.set_configuration
   
        selection_file = open(ifc_properties.my_selectionload)
        selection_configuration = json.load(selection_file)

        custom_collection.items.clear()

        for property_name_from_json, property_value_from_json in selection_configuration.items():
            if property_name_from_json.startswith('my_ifccustomproperty'):
                set_configuration(context,property_name=property_value_from_json)
                
            if not hasattr(ifc_properties, property_name_from_json):
                continue  
            setattr(ifc_properties, property_name_from_json, property_value_from_json)
            
        selection_file.close()

        return {"FINISHED"} 
    
class ClearProperties(bpy.types.Operator):

    bl_idname = "clear.clear_properties"
    bl_label = "Clear Properties"

    def execute(self, context):

        custom_collection = context.scene.custom_collection
        custom_collection.items.clear()

        return {"FINISHED"}
    
class ClearSelection(bpy.types.Operator):

    bl_idname = "save.clear_selection"
    bl_label = "Clear All"

    def execute(self, context):

        ifc_properties = context.scene.ifc_properties
        custom_collection = context.scene.custom_collection
        exlude_list = ['my_spreadsheetfile','ods_or_xlsx','my_selectionload']

        for my_ifcproperty in ifc_properties.__annotations__.keys():
            if my_ifcproperty not in exlude_list:
                setattr(ifc_properties, my_ifcproperty, False)

        custom_collection.items.clear()
       
        return {"FINISHED"} 

        
def register():
    bpy.utils.register_class(ExportToSpreadSheet)
    bpy.utils.register_class(FilterIFCElements)
    bpy.utils.register_class(UnhideIFCElements)
    bpy.utils.register_class(CustomCollectionActions)
    bpy.utils.register_class(SaveAndLoadSelection)
    bpy.utils.register_class(ConfirmSelection)
    bpy.utils.register_class(ClearSelection)
    bpy.utils.register_class(ClearProperties)

    

def unregister():
    bpy.utils.unregister_class(ExportToSpreadSheet)
    bpy.utils.unregister_class(FilterIFCElements)
    bpy.utils.unregister_class(UnhideIFCElements)
    bpy.utils.unregister_class(CustomCollectionActions)
    bpy.utils.unregister_class(SaveAndLoadSelection)
    bpy.utils.unregister_class(ConfirmSelection)
    bpy.utils.unregister_class(ClearSelection)
    bpy.utils.unregister_class(ClearProperties)
